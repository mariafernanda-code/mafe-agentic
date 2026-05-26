"""Generador de Excels."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F3A68")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="F0EEE7")
TOTAL_FONT = Font(bold=True, color="2A2A2A")
THIN = Side(border_style="thin", color="D9D6CC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _autosize(ws, headers, rows):
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in rows:
            if col_idx - 1 < len(r):
                v = r[col_idx - 1]
                if v is not None:
                    max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)


def build(spec: dict, output_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_spec in spec["sheets"]:
        name = sheet_spec["name"][:31]
        headers = sheet_spec.get("headers", [])
        rows = sheet_spec.get("rows", [])
        totals_row = sheet_spec.get("totals_row", False)

        ws = wb.create_sheet(name)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL; cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = BORDER
                if _is_num(value):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        if totals_row and rows:
            total_row_idx = len(rows) + 2
            tc = ws.cell(row=total_row_idx, column=1, value="TOTAL")
            tc.font = TOTAL_FONT; tc.fill = TOTAL_FILL; tc.border = BORDER

            for c_idx in range(2, len(headers) + 1):
                col_letter = get_column_letter(c_idx)
                first_val = rows[0][c_idx - 1] if c_idx - 1 < len(rows[0]) else None
                if _is_num(first_val):
                    formula = f"=SUM({col_letter}2:{col_letter}{len(rows) + 1})"
                    cell = ws.cell(row=total_row_idx, column=c_idx, value=formula)
                    cell.number_format = "#,##0.00"
                else:
                    cell = ws.cell(row=total_row_idx, column=c_idx, value="")
                cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = BORDER
                cell.alignment = Alignment(horizontal="right")

        ws.freeze_panes = "A2"
        _autosize(ws, headers, rows)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
